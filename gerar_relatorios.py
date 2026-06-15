#!/usr/bin/env python3
"""
Gerador de Relatórios de Fechamento por Unidade — 360 Suítes
Uso: python3 gerar_relatorios.py --input FECHAMENTO.xlsx --output ./output-pdfs [--partner ajr] [--mes 5] [--ano 2026]
"""

import sys
import json
import zipfile
import xml.etree.ElementTree as ET
import argparse
from pathlib import Path
from collections import defaultdict

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


# ─── Constantes ───────────────────────────────────────────────────────────────

MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

COR_PRIMARIA    = colors.HexColor("#1A1A2E")
COR_SECUNDARIA  = colors.HexColor("#16213E")
COR_DESTAQUE    = colors.HexColor("#0F3460")
COR_SUBTOTAIS   = colors.HexColor("#1565C0")   # azul médio — PLC, PLCLC, PLCLCLA%
COR_VERDE       = colors.HexColor("#2D6A4F")
COR_VERMELHO    = colors.HexColor("#C1121F")
COR_CINZA_CLARO = colors.HexColor("#F0F4F8")
COR_CINZA       = colors.HexColor("#CCCCCC")
COR_BRANCO      = colors.white
COR_TEXTO       = colors.HexColor("#333333")


# ─── Leitor de XLSX raw (preserva números formatados como data) ───────────────

def col_to_idx(col: str) -> int:
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def _parse_sheet(ws_tree, shared: list[str]) -> list[dict]:
    ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    all_rows = ws_tree.findall(".//ns:row", ns)
    if not all_rows:
        return []

    # Detecta linha de cabeçalho: linha com mais células string (≥ 5)
    header_row_idx = 0
    best_str_count = 0
    for i, row in enumerate(all_rows[:10]):  # verifica só as primeiras 10 linhas
        cells = row.findall("ns:c", ns)
        str_count = sum(1 for c in cells if c.get("t") == "s")
        if str_count > best_str_count:
            best_str_count = str_count
            header_row_idx = i

    headers: dict[int, str] = {}
    for cell in all_rows[header_row_idx].findall("ns:c", ns):
        ref = cell.get("r", "")
        col_str = "".join(c for c in ref if c.isalpha())
        idx = col_to_idx(col_str)
        t = cell.get("t", "")
        v = cell.findtext("ns:v", namespaces=ns)
        if t == "s" and v is not None:
            headers[idx] = shared[int(v)]
        elif v:
            headers[idx] = v

    rows_data = []
    for row in all_rows[header_row_idx + 1:]:
        row_dict: dict = {}
        for cell in row.findall("ns:c", ns):
            ref = cell.get("r", "")
            col_str = "".join(c for c in ref if c.isalpha())
            idx = col_to_idx(col_str)
            header = headers.get(idx)
            if header is None:
                continue
            t = cell.get("t", "")
            v = cell.findtext("ns:v", namespaces=ns)
            if t == "s" and v is not None:
                row_dict[header] = shared[int(v)]
            elif v is not None:
                try:
                    row_dict[header] = float(v)
                except Exception:
                    row_dict[header] = v
            else:
                row_dict[header] = None
        rows_data.append(row_dict)

    return rows_data


def read_xlsx_sheets(path: str) -> dict[str, list[dict]]:
    """Lê todas as abas do XLSX, preservando valores numéricos."""
    with zipfile.ZipFile(path) as z:
        # Shared strings
        with z.open("xl/sharedStrings.xml") as f:
            ss_tree = ET.fromstring(f.read())

        # Mapa nome_aba → arquivo
        ns_wb = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                 "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
        with z.open("xl/workbook.xml") as f:
            wb_tree = ET.fromstring(f.read())
        with z.open("xl/_rels/workbook.xml.rels") as f:
            rels_tree = ET.fromstring(f.read())

        ns_r = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rels = {rel.get("Id"): rel.get("Target") for rel in rels_tree.findall("r:Relationship", ns_r)}

        ns_main = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                   "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
        sheets_map = {}
        for sheet in wb_tree.findall(".//ns:sheet", ns_main):
            name = sheet.get("name", "")
            rid = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            target = rels.get(rid, "")
            if target:
                sheets_map[name] = f"xl/{target}" if not target.startswith("xl/") else target

        shared = []
        for si in ss_tree.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"):
            texts = si.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
            shared.append("".join(t.text or "" for t in texts))

        result = {}
        for sheet_name, ws_path in sheets_map.items():
            try:
                with z.open(ws_path) as f:
                    ws_tree = ET.fromstring(f.read())
                result[sheet_name] = _parse_sheet(ws_tree, shared)
            except Exception as e:
                print(f"  ⚠️  Erro ao ler aba '{sheet_name}': {e}")

    return result


# ─── Agregação de reservas por unidade ────────────────────────────────────────

def agregar_reservas(reservas_rows: list[dict]) -> dict[str, dict]:
    """
    Retorna mapa unit_code → {noites: int, reservas: int, por_plataforma: {plat: noites}}.
    Considera apenas linhas com NOITES MÊS > 0.
    """
    resultado: dict[str, dict] = {}
    for row in reservas_rows:
        unit = str(row.get("LISTING'S NICKNAME") or "").strip()
        noites_mes = float(row.get("NOITES MÊS") or 0)
        if not unit or noites_mes <= 0:
            continue
        plataforma = str(row.get("PLATFORM") or row.get("Source") or "—").strip()
        if unit not in resultado:
            resultado[unit] = {"noites": 0, "reservas": 0, "por_plataforma": defaultdict(int)}
        resultado[unit]["noites"] += noites_mes
        resultado[unit]["reservas"] += 1
        resultado[unit]["por_plataforma"][plataforma] += noites_mes
    return resultado


# ─── Helpers ──────────────────────────────────────────────────────────────────

def fmt(value, prefix="R$ ") -> str:
    if value is None or (isinstance(value, float) and value != value):
        return "—"
    try:
        v = float(value)
        neg = v < 0
        s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"-{prefix}{s}" if neg else f"{prefix}{s}"
    except Exception:
        return str(value)


def n(value) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def load_predios(predios_path: str) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(predios_path)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        unit_idx = headers.index("unit") if "unit" in headers else 0
        name_idx = headers.index("property_name") if "property_name" in headers else -1
        resultado = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[unit_idx] and name_idx >= 0 and row[name_idx]:
                unit = str(row[unit_idx]).strip()
                sigla = "".join(c for c in unit if not c.isdigit()).strip().upper()
                if sigla and sigla not in resultado:
                    resultado[sigla] = str(row[name_idx]).strip()
        return resultado
    except Exception:
        return {}


def get_nome_predio(unit: str, predios: dict) -> str:
    sigla = "".join(c for c in unit if not c.isdigit()).strip().upper()
    return predios.get(sigla, f"Unidade {unit}")


# ─── Componentes visuais ──────────────────────────────────────────────────────

def make_styles(doc):
    base = getSampleStyleSheet()

    def s(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    W = doc.width
    return W, {
        "title":       s("title",   fontSize=18, textColor=COR_BRANCO, alignment=TA_CENTER, fontName="Helvetica-Bold"),
        "subtitle":    s("subt",    fontSize=11, textColor=COR_CINZA,  alignment=TA_CENTER, fontName="Helvetica"),
        "owner_name":  s("own",     fontSize=12, textColor=COR_PRIMARIA, fontName="Helvetica-Bold", alignment=TA_CENTER),
        "section":     s("sec",     fontSize=9,  textColor=COR_BRANCO, fontName="Helvetica-Bold"),
        "label":       s("lbl",     fontSize=8,  textColor=COR_TEXTO,  fontName="Helvetica"),
        "label_bold":  s("lblb",    fontSize=8,  textColor=COR_TEXTO,  fontName="Helvetica-Bold"),
        "value":       s("val",     fontSize=8,  textColor=COR_TEXTO,  fontName="Helvetica-Bold", alignment=TA_RIGHT),
        "sub_lbl":     s("sublbl",  fontSize=9,  textColor=COR_BRANCO, fontName="Helvetica-Bold"),
        "sub_val":     s("subval",  fontSize=9,  textColor=COR_BRANCO, fontName="Helvetica-Bold", alignment=TA_RIGHT),
        "total_lbl":   s("totlbl",  fontSize=9,  textColor=COR_BRANCO, fontName="Helvetica-Bold"),
        "total_val":   s("totval",  fontSize=9,  textColor=COR_BRANCO, fontName="Helvetica-Bold", alignment=TA_RIGHT),
        "note":        s("note",    fontSize=7,  textColor=colors.HexColor("#555555"), fontName="Helvetica-Oblique"),
        "footer":      s("foot",    fontSize=7,  textColor=colors.HexColor("#888888"), alignment=TA_CENTER),
        "plat_lbl":    s("plat",    fontSize=7,  textColor=COR_TEXTO,  fontName="Helvetica"),
        "plat_val":    s("platv",   fontSize=7,  textColor=COR_TEXTO,  fontName="Helvetica-Bold", alignment=TA_RIGHT),
    }


def section_header(title: str, W: float, st: dict, cor=None):
    cor = cor or COR_DESTAQUE
    return Table(
        [[Paragraph(title, st["section"])]],
        colWidths=[W],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), cor),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ]),
    )


def row_normal(label: str, value: str, W: float, st: dict, indent: int = 0):
    """Linha simples: label à esq, valor à dir."""
    pad = indent * 8
    return Table(
        [[Paragraph(label, st["label"]), Paragraph(value, st["value"])]],
        colWidths=[W * 0.65, W * 0.35],
        style=TableStyle([
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8 + pad),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.3, COR_CINZA),
        ]),
    )


def row_subtotal(label: str, value: str, W: float, st: dict):
    """Linha de subtotal destacada em azul escuro (PLC, PLCLC, PLCLCLA%)."""
    return Table(
        [[Paragraph(label, st["sub_lbl"]), Paragraph(value, st["sub_val"])]],
        colWidths=[W * 0.65, W * 0.35],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), COR_SUBTOTAIS),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, COR_BRANCO),
        ]),
    )


def row_resultado(label: str, value: str, W: float, st: dict):
    """Linha cinza claro — resultado intermediário após despesas."""
    return Table(
        [[Paragraph(label, st["label_bold"]), Paragraph(value, st["value"])]],
        colWidths=[W * 0.65, W * 0.35],
        style=TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), COR_CINZA_CLARO),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.3, COR_CINZA),
        ]),
    )


# ─── PDF por unidade ──────────────────────────────────────────────────────────

def build_pdf(unit_data: dict, output_path: str, mes: int, ano: int,
              nome_predio: str, owner_name: str, reservas_info: dict | None):
    unit     = unit_data.get("Unidades", "")
    dealtype = str(unit_data.get("Dealtype", "ADM")).upper()

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    W, st = make_styles(doc)

    # ── Valores ──
    fat_bruto     = n(unit_data.get("Fat_mês"))
    comm_ota      = n(unit_data.get("Comm OTA"))
    comm_360      = n(unit_data.get("Comm 360"))
    comm_partner  = n(unit_data.get("CommPartner"))
    plc           = n(unit_data.get("PLC"))
    limpeza       = n(unit_data.get("Limpeza Ajustada"))
    plclc         = n(unit_data.get("PLCLC"))
    tx_partner    = n(unit_data.get("% partner"))
    tx_360        = n(unit_data.get("3.6"))
    plclcla       = n(unit_data.get("PLCLCLA%"))
    energia       = n(unit_data.get("Energia Elétrica - Apto"))
    condominio    = n(unit_data.get("Condomínio - Apto"))
    internet      = n(unit_data.get("Internet - Apto"))
    iptu          = n(unit_data.get("IPTU - Apto"))
    seguros       = n(unit_data.get("Seguros"))
    housi_pay     = n(unit_data.get("Housi Pay"))
    implantacao   = n(unit_data.get("Implantação - Compras e Serviços"))
    despesas_desc = n(unit_data.get("Despesas - Desconto no Repasse"))
    manutencoes   = n(unit_data.get("Manutenções e Serviços Diários - descontos"))
    reembolso     = n(unit_data.get("Reembolso despesas"))
    resultado     = n(unit_data.get("PLCLCA%LDESP"))
    fixo          = n(unit_data.get("Fixpo ajustado") or unit_data.get("Fixo ajustado") or unit_data.get("Fixo"))
    min_gar       = n(unit_data.get("Min ajustado") or unit_data.get("Min"))
    repasse       = n(unit_data.get("Repasse Cliene"))
    checkins      = int(n(unit_data.get("Checkin")))
    checkouts     = int(n(unit_data.get("Checkout")))

    perc_partner  = n(unit_data.get("Taxa Partner")) * 100
    perc_360      = n(unit_data.get("Taxa 360")) * 100

    story = []

    # ── Cabeçalho ──
    story.append(Table([[Paragraph("360 Suítes", st["title"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_PRIMARIA),
                          ("TOPPADDING",(0,0),(-1,-1),14),("BOTTOMPADDING",(0,0),(-1,-1),8),
                          ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10)])))
    story.append(Table([[Paragraph(f"Relatório de Fechamento — {MESES[mes]}/{ano}", st["subtitle"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_SECUNDARIA),
                          ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                          ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10)])))
    story.append(Spacer(1, 0.3*cm))

    # ── Identificação ──
    story.append(section_header("📋  IDENTIFICAÇÃO DA UNIDADE", W, st))
    id_rows = [
        ["Unidade",        unit,                  "Proprietário",         owner_name],
        ["Empreendimento", nome_predio,            "Tipo de Contrato",     dealtype],
        ["Período",        f"{MESES[mes]}/{ano}", "Check-ins / Check-outs", f"{checkins} / {checkouts}"],
    ]
    story.append(Table(id_rows, colWidths=[W*0.18, W*0.37, W*0.20, W*0.25],
        style=TableStyle([
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("TEXTCOLOR",(0,0),(0,-1),COR_DESTAQUE),("TEXTCOLOR",(2,0),(2,-1),COR_DESTAQUE),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[COR_CINZA_CLARO, COR_BRANCO]),
            ("LINEBELOW",(0,0),(-1,-1),0.3,COR_CINZA),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ])))
    story.append(Spacer(1, 0.3*cm))

    # ── Ocupação (da aba RESERVAS) ──
    ri = reservas_info or {}
    noites_total = int(ri.get("noites", 0))
    n_reservas   = int(ri.get("reservas", 0))
    por_plat     = ri.get("por_plataforma", {})
    dias_mes     = 31  # Maio tem 31 dias
    tx_ocp       = f"{(noites_total / dias_mes * 100):.1f}%" if noites_total else "—"

    story.append(section_header("🌙  OCUPAÇÃO DO MÊS", W, st))
    ocp_rows = [
        ["Noites ocupadas no mês",  f"{noites_total} de {dias_mes}",  "Taxa de ocupação",    tx_ocp],
        ["Número de reservas",       str(n_reservas),                  "Dias disponíveis",    str(dias_mes - noites_total)],
    ]
    story.append(Table(ocp_rows, colWidths=[W*0.22, W*0.28, W*0.22, W*0.28],
        style=TableStyle([
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("TEXTCOLOR",(0,0),(0,-1),COR_DESTAQUE),("TEXTCOLOR",(2,0),(2,-1),COR_DESTAQUE),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[COR_CINZA_CLARO, COR_BRANCO]),
            ("LINEBELOW",(0,0),(-1,-1),0.3,COR_CINZA),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ])))

    if por_plat:
        plat_rows = [[Paragraph(f"• {p}", st["plat_lbl"]),
                      Paragraph(f"{int(n)} noites", st["plat_val"])]
                     for p, n in sorted(por_plat.items(), key=lambda x: -x[1])]
        story.append(Table(plat_rows, colWidths=[W*0.65, W*0.35],
            style=TableStyle([
                ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
                ("LEFTPADDING",(0,0),(-1,-1),16),("RIGHTPADDING",(0,0),(-1,-1),8),
                ("LINEBELOW",(0,0),(-1,-1),0.2,COR_CINZA),
            ])))
    story.append(Spacer(1, 0.3*cm))

    # ── Receita ──
    story.append(section_header("💰  RECEITA", W, st))
    story.append(row_normal("Faturamento Bruto (total gerado pelas reservas)", fmt(fat_bruto), W, st))
    story.append(row_normal("(-) Comissão OTA (taxa cobrada pela plataforma)", fmt(-comm_ota), W, st, indent=1))
    if comm_360:
        story.append(row_normal("(-) Comissão 360", fmt(-comm_360), W, st, indent=1))
    if comm_partner:
        story.append(row_normal("(-) Comissão Partner", fmt(-comm_partner), W, st, indent=1))

    story.append(row_subtotal(
        "= PLC — Preço Líquido de Comissão  (faturamento após deduzir comissões)",
        fmt(plc), W, st))

    story.append(row_normal("(-) Limpeza (receita de limpeza repassada ao serviço)", fmt(-limpeza), W, st, indent=1))

    story.append(row_subtotal(
        "= PLCLC — Preço Líq. de Comissão e Limpeza  (base de cálculo das taxas)",
        fmt(plclc), W, st))
    story.append(Spacer(1, 0.3*cm))

    # ── Taxas de Gestão (ADM) ──
    if dealtype == "ADM":
        story.append(section_header("📊  TAXAS DE GESTÃO", W, st))
        if tx_partner:
            story.append(row_normal(f"(-) Taxa Partner ({perc_partner:.0f}% sobre PLCLC)", fmt(-tx_partner), W, st, indent=1))
        if tx_360:
            story.append(row_normal(f"(-) Taxa 360 ({perc_360:.0f}% sobre PLCLC)", fmt(-tx_360), W, st, indent=1))
        story.append(row_subtotal(
            "= PLCLCLA% — Preço Líq. após Comissões, Limpeza e Adm.  (base para dedução de despesas)",
            fmt(plclcla), W, st))
        story.append(Spacer(1, 0.3*cm))
    else:
        plclcla = plclc

    # ── Despesas ──
    despesas = [
        ("Energia Elétrica",                    energia),
        ("Condomínio",                           condominio),
        ("Internet",                             internet),
        ("IPTU",                                 iptu),
        ("Seguros",                              seguros),
        ("Housi Pay",                            housi_pay),
        ("Implantação / Compras e Serviços",     implantacao),
        ("Manutenções e Serviços Diários",       manutencoes),
        ("Despesas — Desconto no Repasse",       despesas_desc),
    ]
    despesas_existentes = [(l, v) for l, v in despesas if v]
    if despesas_existentes or reembolso:
        story.append(section_header("🏠  DESPESAS", W, st))
        for label, val in despesas_existentes:
            story.append(row_normal(f"(-) {label}", fmt(-val), W, st, indent=1))
        if reembolso:
            story.append(row_normal("(+) Reembolso de Despesas", fmt(reembolso), W, st, indent=1))
        story.append(row_resultado("= Resultado após Despesas", fmt(resultado), W, st))
        story.append(Spacer(1, 0.3*cm))
    else:
        # Sem despesas — mostra linha informativa e usa plclcla como resultado
        story.append(section_header("🏠  DESPESAS", W, st))
        story.append(row_normal("Nenhuma despesa registrada no período", "—", W, st))
        story.append(Spacer(1, 0.3*cm))

    # ── Repasse ──
    story.append(section_header("✅  REPASSE AO PROPRIETÁRIO", W, st, cor=COR_VERDE))
    if dealtype == "FIX":
        if fixo:
            story.append(row_normal("Valor Fixo Contratado", fmt(fixo), W, st))
        if min_gar:
            story.append(row_normal("Mínimo Garantido", fmt(min_gar), W, st))

    repasse_cor = COR_VERDE if repasse >= 0 else COR_VERMELHO
    story.append(Table(
        [[Paragraph("REPASSE LÍQUIDO AO PROPRIETÁRIO", st["total_lbl"]),
          Paragraph(fmt(repasse), st["total_val"])]],
        colWidths=[W*0.65, W*0.35],
        style=TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),repasse_cor),
            ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),10),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
        ])))
    story.append(Spacer(1, 0.4*cm))

    # ── Rodapé ──
    story.append(HRFlowable(width=W, thickness=0.5, color=COR_CINZA))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"360 Suítes | proprietarios@360suites.com.br | www.360suites.com.br<br/>"
        f"Documento gerado automaticamente — {MESES[mes]}/{ano}. Dúvidas? Entre em contato com nossa equipe.",
        st["footer"]))

    doc.build(story)


# ─── PDF resumo por proprietário ──────────────────────────────────────────────

def build_summary_pdf(owner_name: str, units_data: list[dict], output_path: str,
                      mes: int, ano: int, predios: dict, reservas_map: dict):
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    W, st = make_styles(doc)
    story = []

    story.append(Table([[Paragraph("360 Suítes", st["title"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_PRIMARIA),
                          ("TOPPADDING",(0,0),(-1,-1),14),("BOTTOMPADDING",(0,0),(-1,-1),8),
                          ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10)])))
    story.append(Table([[Paragraph(f"Resumo de Fechamento — {MESES[mes]}/{ano}", st["subtitle"])]],
        colWidths=[W],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_SECUNDARIA),
                          ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                          ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10)])))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Proprietário: {owner_name}", st["owner_name"]))
    story.append(Spacer(1, 0.3*cm))

    total_repasse = 0.0

    for ud in units_data:
        unit        = ud.get("Unidades", "")
        nome_predio = get_nome_predio(unit, predios)
        dealtype    = str(ud.get("Dealtype", "ADM")).upper()
        repasse     = n(ud.get("Repasse Cliene"))
        fat_bruto   = n(ud.get("Fat_mês"))
        plclc       = n(ud.get("PLCLC"))
        resultado   = n(ud.get("PLCLCA%LDESP"))
        ri          = reservas_map.get(unit, {})
        noites      = int(ri.get("noites", 0))
        n_res       = int(ri.get("reservas", 0))
        total_repasse += repasse

        unit_block = []
        unit_block.append(Table(
            [[Paragraph(f"🏢  {unit} — {nome_predio}  ({dealtype})", st["section"])]],
            colWidths=[W],
            style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_DESTAQUE),
                              ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                              ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8)])))

        summary_data = [
            ["Faturamento Bruto", fmt(fat_bruto),     "PLCLC",          fmt(plclc)],
            [f"Noites ocupadas",  f"{noites} noites", "Resultado (líq.)", fmt(resultado)],
            ["Reservas no mês",   f"{n_res}",          "Tipo contrato",  dealtype],
        ]
        unit_block.append(Table(summary_data, colWidths=[W*0.20, W*0.30, W*0.22, W*0.28],
            style=TableStyle([
                ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),8),
                ("TEXTCOLOR",(0,0),(0,-1),COR_DESTAQUE),("TEXTCOLOR",(2,0),(2,-1),COR_DESTAQUE),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
                ("ROWBACKGROUNDS",(0,0),(-1,-1),[COR_CINZA_CLARO, COR_BRANCO]),
                ("LINEBELOW",(0,0),(-1,-1),0.3,COR_CINZA),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ])))

        repasse_cor = COR_VERDE if repasse >= 0 else COR_VERMELHO
        unit_block.append(Table(
            [[Paragraph(f"REPASSE — {unit}", st["total_lbl"]),
              Paragraph(fmt(repasse), st["total_val"])]],
            colWidths=[W*0.65, W*0.35],
            style=TableStyle([("BACKGROUND",(0,0),(-1,-1),repasse_cor),
                              ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                              ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8)])))
        unit_block.append(Spacer(1, 0.4*cm))
        story.append(KeepTogether(unit_block))

    story.append(Table(
        [[Paragraph("TOTAL CONSOLIDADO — REPASSE AO PROPRIETÁRIO", st["total_lbl"]),
          Paragraph(fmt(total_repasse), st["total_val"])]],
        colWidths=[W*0.65, W*0.35],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),COR_PRIMARIA),
                          ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
                          ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8)])))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width=W, thickness=0.5, color=COR_CINZA))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"360 Suítes | proprietarios@360suites.com.br | www.360suites.com.br<br/>"
        f"Documento gerado automaticamente — {MESES[mes]}/{ano}.",
        st["footer"]))

    doc.build(story)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gerador de Relatórios de Fechamento — 360 Suítes")
    parser.add_argument("--input",          required=True,  help="Caminho do fechamento.xlsx")
    parser.add_argument("--output",         default="output-fechamento-pdfs")
    parser.add_argument("--partner",        default="ajr",  help="Filtro parceiro (padrão: ajr | 'all' para todos)")
    parser.add_argument("--mes",            type=int, default=5)
    parser.add_argument("--ano",            type=int, default=2026)
    parser.add_argument("--predios",        default=None)
    parser.add_argument("--proprietarios",  default=None)
    parser.add_argument("--unit",           default=None, help="Gerar só esta unidade (teste)")
    args = parser.parse_args()

    script_dir   = Path(__file__).parent
    predios_path = args.predios or str(script_dir / "predios.xlsx")
    props_path   = args.proprietarios or str(script_dir / "proprietarios.json")

    print(f"📂 Lendo {args.input} ...")
    sheets = read_xlsx_sheets(args.input)

    # Aba UNIDADES
    unidades_key = next((k for k in sheets if "unidade" in k.lower()), None)
    if not unidades_key:
        print("❌ Aba UNIDADES não encontrada."); sys.exit(1)
    rows = [r for r in sheets[unidades_key] if r.get("Unidades")]

    if args.partner and args.partner.lower() != "all":
        rows = [r for r in rows if r.get("Partner") and args.partner.lower() in str(r["Partner"]).lower()]
    print(f"📊 {len(rows)} unidades (parceiro: {args.partner})")

    if args.unit:
        rows = [r for r in rows if r.get("Unidades", "").upper() == args.unit.upper()]
        if not rows:
            print(f"❌ Unidade '{args.unit}' não encontrada."); sys.exit(1)

    # Aba RESERVAS
    reservas_key = next((k for k in sheets if "reserva" in k.lower()), None)
    reservas_map: dict[str, dict] = {}
    if reservas_key:
        reservas_map = agregar_reservas(sheets[reservas_key])
        print(f"🌙 Reservas carregadas: {len(reservas_map)} unidades com dados")
    else:
        print("⚠️  Aba RESERVAS não encontrada — noites/reservas não serão exibidas")

    predios   = load_predios(predios_path) if Path(predios_path).exists() else {}
    print(f"🏢 {len(predios)} prédios carregados")

    owner_map: dict[str, str] = {}
    if Path(props_path).exists():
        for p in json.loads(Path(props_path).read_text("utf-8")):
            for u in p.get("unidades", []):
                owner_map[u] = p.get("nome", "")

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"📁 Saída: {out_dir.resolve()}\n")

    ok = erros = 0
    owners_units: dict[str, list[dict]] = {}

    for ud in rows:
        unit            = ud.get("Unidades", "")
        owner_from_xlsx = str(ud.get("Owner", "")).strip()
        owner_name      = owner_map.get(unit, owner_from_xlsx)
        nome_predio     = get_nome_predio(unit, predios)
        ri              = reservas_map.get(unit)

        owners_units.setdefault(owner_from_xlsx, []).append(ud)

        pdf_path = str(out_dir / f"{unit} - {owner_name}.pdf")
        try:
            build_pdf(ud, pdf_path, args.mes, args.ano, nome_predio, owner_name, ri)
            print(f"  ✅ {unit} - {owner_name}.pdf")
            ok += 1
        except Exception as e:
            print(f"  ❌ {unit}: {e}")
            erros += 1

    print(f"\n📋 Gerando resumos por proprietário...")
    for owner_from_xlsx, units in owners_units.items():
        owner_name   = owner_map.get(units[0].get("Unidades", ""), owner_from_xlsx)
        summary_path = str(out_dir / f"Resumo - {owner_name}.pdf")
        try:
            build_summary_pdf(owner_name, units, summary_path, args.mes, args.ano, predios, reservas_map)
            print(f"  📄 Resumo - {owner_name}.pdf ({len(units)} unidade(s))")
        except Exception as e:
            print(f"  ❌ Resumo {owner_name}: {e}")

    print(f"\n✅ {ok} PDFs gerados, {erros} erros")
    print(f"📁 {out_dir.resolve()}")


if __name__ == "__main__":
    main()
